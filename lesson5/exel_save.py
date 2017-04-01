from open_name import get_open_name
from openpyxl import Workbook

def exel_write():
	data = get_open_name('http://api.data.mos.ru/v1/datasets/2009/rows')
	workbook = Workbook()
	workseet = workbook.active
	workseet.title = 'Имена_Москвы'
	name_column = {}
	t_col=1
	for column in data[0]['Cells']:
		name_column[column] = t_col
		workseet.cell(row=1,column=t_col).value  =  column
		t_col += 1
	t_row=2
	for item in data:
		for column in item['Cells']:
			workseet.cell(row=t_row,column=name_column[column]).value  =  item['Cells'][column]
		t_row += 1
	workbook.save('NameMoscow.xlsx')


	    
	return t_row
if __name__ == '__main__':
	print(exel_write())

